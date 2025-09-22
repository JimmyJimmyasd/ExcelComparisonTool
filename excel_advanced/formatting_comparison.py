# Excel Formatting Comparison Module
# Comprehensive analysis and comparison of Excel cell formatting, conditional formatting, and structure

import pandas as pd
import numpy as np
from typing import Dict, List, Optional, Tuple, Any, Union
from datetime import datetime
import json
from collections import defaultdict, Counter

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Fill, Border, Alignment, Protection
    from openpyxl.formatting.rule import Rule, ColorScaleRule, DataBarRule, IconSetRule
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not available. Formatting comparison will be disabled.")

class FormattingComparator:
    """
    Advanced Excel formatting comparison and analysis tool.
    Compares cell formatting, conditional formatting, and sheet structure.
    """
    
    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for formatting comparison")
        
        # Color mappings for better readability
        self.color_mappings = {
            'theme_colors': {
                0: 'White', 1: 'Black', 2: 'Blue', 3: 'Red', 4: 'Green', 
                5: 'Yellow', 6: 'Magenta', 7: 'Cyan'
            }
        }
    
    def compare_excel_formatting(self, file_a: Any, file_b: Any, 
                                sheet_name_a: Optional[str] = None, 
                                sheet_name_b: Optional[str] = None) -> Dict[str, Any]:
        """
        Compare formatting between two Excel files or sheets.
        
        Args:
            file_a: First Excel file (file path or file-like object)
            file_b: Second Excel file (file path or file-like object)
            sheet_name_a: Sheet name in first file (default: first sheet)
            sheet_name_b: Sheet name in second file (default: first sheet)
        
        Returns:
            Dictionary containing comprehensive formatting comparison results
        """
        try:
            # Load workbooks
            wb_a = load_workbook(file_a, data_only=False)
            wb_b = load_workbook(file_b, data_only=False)
            
            # Get sheets
            sheet_a = wb_a[sheet_name_a] if sheet_name_a else wb_a.active
            sheet_b = wb_b[sheet_name_b] if sheet_name_b else wb_b.active
            
            # Perform comprehensive formatting comparison
            results = {
                'status': 'success',
                'timestamp': datetime.now().isoformat(),
                'files_compared': {
                    'file_a': getattr(file_a, 'name', 'File A'),
                    'file_b': getattr(file_b, 'name', 'File B'),
                    'sheet_a': sheet_a.title,
                    'sheet_b': sheet_b.title
                },
                'cell_formatting': self._compare_cell_formatting(sheet_a, sheet_b),
                'conditional_formatting': self._compare_conditional_formatting(wb_a, wb_b, sheet_a, sheet_b),
                'structure_comparison': self._compare_sheet_structure(sheet_a, sheet_b),
                'summary_statistics': self._calculate_formatting_statistics(sheet_a, sheet_b),
                'formatting_diff_report': self._generate_formatting_diff_report(sheet_a, sheet_b)
            }
            
            return results
            
        except Exception as e:
            return {
                'status': 'error',
                'error': str(e),
                'error_type': type(e).__name__,
                'timestamp': datetime.now().isoformat()
            }
    
    def _compare_cell_formatting(self, sheet_a, sheet_b) -> Dict[str, Any]:
        """Compare cell-level formatting between two sheets."""
        
        # Get all cells with data or formatting in both sheets
        cells_a = self._get_formatted_cells(sheet_a)
        cells_b = self._get_formatted_cells(sheet_b)
        
        # Find all unique cell coordinates
        all_coords = set(cells_a.keys()) | set(cells_b.keys())
        
        formatting_differences = {}
        format_categories = {
            'font_differences': [],
            'fill_differences': [],
            'border_differences': [],
            'alignment_differences': [],
            'number_format_differences': [],
            'protection_differences': []
        }
        
        for coord in all_coords:
            cell_a = cells_a.get(coord)
            cell_b = cells_b.get(coord)
            
            diff = self._compare_cell_formats(coord, cell_a, cell_b)
            if diff['has_differences']:
                formatting_differences[coord] = diff
                
                # Categorize differences
                for category, diffs in diff['differences'].items():
                    if diffs:
                        format_categories[f"{category}_differences"].append({
                            'cell': coord,
                            'changes': diffs
                        })
        
        return {
            'total_cells_compared': len(all_coords),
            'cells_with_differences': len(formatting_differences),
            'difference_percentage': (len(formatting_differences) / len(all_coords) * 100) if all_coords else 0,
            'formatting_differences': formatting_differences,
            'categorized_differences': format_categories,
            'summary': self._summarize_formatting_differences(format_categories)
        }
    
    def _get_formatted_cells(self, sheet) -> Dict[str, Any]:
        """Extract all cells that have data or formatting."""
        formatted_cells = {}
        
        # Iterate through all cells in the used range
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None or self._has_custom_formatting(cell):
                    coord = f"{cell.column_letter}{cell.row}"
                    formatted_cells[coord] = cell
        
        return formatted_cells
    
    def _has_custom_formatting(self, cell) -> bool:
        """Check if a cell has custom formatting beyond defaults."""
        # Check if any formatting property differs from default
        default_font = Font()
        default_fill = Fill()
        default_border = Border()
        default_alignment = Alignment()
        
        return (
            cell.font != default_font or
            cell.fill != default_fill or
            cell.border != default_border or
            cell.alignment != default_alignment or
            cell.number_format != 'General'
        )
    
    def _compare_cell_formats(self, coord: str, cell_a, cell_b) -> Dict[str, Any]:
        """Compare formatting between two cells."""
        differences = {
            'font': [],
            'fill': [],
            'border': [],
            'alignment': [],
            'number_format': [],
            'protection': []
        }
        
        has_differences = False
        
        # Compare fonts
        if cell_a and cell_b:
            font_diff = self._compare_fonts(cell_a.font, cell_b.font)
            if font_diff:
                differences['font'] = font_diff
                has_differences = True
            
            # Compare fills
            fill_diff = self._compare_fills(cell_a.fill, cell_b.fill)
            if fill_diff:
                differences['fill'] = fill_diff
                has_differences = True
            
            # Compare borders
            border_diff = self._compare_borders(cell_a.border, cell_b.border)
            if border_diff:
                differences['border'] = border_diff
                has_differences = True
            
            # Compare alignment
            alignment_diff = self._compare_alignment(cell_a.alignment, cell_b.alignment)
            if alignment_diff:
                differences['alignment'] = alignment_diff
                has_differences = True
            
            # Compare number format
            if cell_a.number_format != cell_b.number_format:
                differences['number_format'] = [
                    f"File A: {cell_a.number_format}",
                    f"File B: {cell_b.number_format}"
                ]
                has_differences = True
            
            # Compare protection
            protection_diff = self._compare_protection(cell_a.protection, cell_b.protection)
            if protection_diff:
                differences['protection'] = protection_diff
                has_differences = True
        
        elif cell_a and not cell_b:
            differences['general'] = ["Cell exists in File A but not in File B"]
            has_differences = True
        elif cell_b and not cell_a:
            differences['general'] = ["Cell exists in File B but not in File A"]
            has_differences = True
        
        return {
            'has_differences': has_differences,
            'differences': differences
        }
    
    def _compare_fonts(self, font_a: Font, font_b: Font) -> List[str]:
        """Compare font properties between two cells."""
        differences = []
        
        if font_a.name != font_b.name:
            differences.append(f"Font name: {font_a.name} → {font_b.name}")
        
        if font_a.size != font_b.size:
            differences.append(f"Font size: {font_a.size} → {font_b.size}")
        
        if font_a.bold != font_b.bold:
            differences.append(f"Bold: {font_a.bold} → {font_b.bold}")
        
        if font_a.italic != font_b.italic:
            differences.append(f"Italic: {font_a.italic} → {font_b.italic}")
        
        if font_a.underline != font_b.underline:
            differences.append(f"Underline: {font_a.underline} → {font_b.underline}")
        
        if font_a.strike != font_b.strike:
            differences.append(f"Strikethrough: {font_a.strike} → {font_b.strike}")
        
        # Compare colors
        color_a = self._get_color_description(font_a.color)
        color_b = self._get_color_description(font_b.color)
        if color_a != color_b:
            differences.append(f"Font color: {color_a} → {color_b}")
        
        return differences
    
    def _compare_fills(self, fill_a: Fill, fill_b: Fill) -> List[str]:
        """Compare fill properties between two cells."""
        differences = []
        
        if fill_a.fill_type != fill_b.fill_type:
            differences.append(f"Fill type: {fill_a.fill_type} → {fill_b.fill_type}")
        
        # Compare foreground colors
        fg_a = self._get_color_description(fill_a.fgColor) if hasattr(fill_a, 'fgColor') else 'None'
        fg_b = self._get_color_description(fill_b.fgColor) if hasattr(fill_b, 'fgColor') else 'None'
        if fg_a != fg_b:
            differences.append(f"Fill color: {fg_a} → {fg_b}")
        
        # Compare background colors for patterns
        if hasattr(fill_a, 'bgColor') and hasattr(fill_b, 'bgColor'):
            bg_a = self._get_color_description(fill_a.bgColor)
            bg_b = self._get_color_description(fill_b.bgColor)
            if bg_a != bg_b:
                differences.append(f"Pattern color: {bg_a} → {bg_b}")
        
        return differences
    
    def _compare_borders(self, border_a: Border, border_b: Border) -> List[str]:
        """Compare border properties between two cells."""
        differences = []
        
        border_sides = ['left', 'right', 'top', 'bottom', 'diagonal']
        
        for side in border_sides:
            side_a = getattr(border_a, side, None)
            side_b = getattr(border_b, side, None)
            
            if side_a and side_b:
                # Compare border styles
                if side_a.style != side_b.style:
                    differences.append(f"{side.title()} border style: {side_a.style} → {side_b.style}")
                
                # Compare border colors
                color_a = self._get_color_description(side_a.color) if side_a.color else 'None'
                color_b = self._get_color_description(side_b.color) if side_b.color else 'None'
                if color_a != color_b:
                    differences.append(f"{side.title()} border color: {color_a} → {color_b}")
            
            elif side_a and not side_b:
                differences.append(f"{side.title()} border removed")
            elif side_b and not side_a:
                differences.append(f"{side.title()} border added: {side_b.style}")
        
        return differences
    
    def _compare_alignment(self, align_a: Alignment, align_b: Alignment) -> List[str]:
        """Compare alignment properties between two cells."""
        differences = []
        
        if align_a.horizontal != align_b.horizontal:
            differences.append(f"Horizontal alignment: {align_a.horizontal} → {align_b.horizontal}")
        
        if align_a.vertical != align_b.vertical:
            differences.append(f"Vertical alignment: {align_a.vertical} → {align_b.vertical}")
        
        if align_a.wrap_text != align_b.wrap_text:
            differences.append(f"Wrap text: {align_a.wrap_text} → {align_b.wrap_text}")
        
        if align_a.shrink_to_fit != align_b.shrink_to_fit:
            differences.append(f"Shrink to fit: {align_a.shrink_to_fit} → {align_b.shrink_to_fit}")
        
        if align_a.indent != align_b.indent:
            differences.append(f"Indent: {align_a.indent} → {align_b.indent}")
        
        if align_a.text_rotation != align_b.text_rotation:
            differences.append(f"Text rotation: {align_a.text_rotation} → {align_b.text_rotation}")
        
        return differences
    
    def _compare_protection(self, prot_a: Protection, prot_b: Protection) -> List[str]:
        """Compare protection properties between two cells."""
        differences = []
        
        if prot_a.locked != prot_b.locked:
            differences.append(f"Locked: {prot_a.locked} → {prot_b.locked}")
        
        if prot_a.hidden != prot_b.hidden:
            differences.append(f"Hidden: {prot_a.hidden} → {prot_b.hidden}")
        
        return differences
    
    def _get_color_description(self, color) -> str:
        """Get a human-readable description of a color."""
        if not color:
            return 'None'
        
        # Handle different color types
        if hasattr(color, 'rgb') and color.rgb:
            return f"RGB({color.rgb})"
        elif hasattr(color, 'theme') and color.theme is not None:
            theme_name = self.color_mappings['theme_colors'].get(color.theme, f"Theme{color.theme}")
            return f"{theme_name}"
        elif hasattr(color, 'indexed') and color.indexed is not None:
            return f"Indexed({color.indexed})"
        else:
            return str(color)
    
    def _compare_conditional_formatting(self, wb_a, wb_b, sheet_a, sheet_b) -> Dict[str, Any]:
        """Compare conditional formatting rules between two sheets."""
        
        # Get conditional formatting rules for both sheets
        cf_rules_a = self._extract_conditional_formatting(sheet_a)
        cf_rules_b = self._extract_conditional_formatting(sheet_b)
        
        return {
            'sheet_a_rules': len(cf_rules_a),
            'sheet_b_rules': len(cf_rules_b),
            'rules_comparison': self._compare_cf_rules(cf_rules_a, cf_rules_b),
            'rules_details_a': cf_rules_a,
            'rules_details_b': cf_rules_b
        }
    
    def _extract_conditional_formatting(self, sheet) -> List[Dict[str, Any]]:
        """Extract conditional formatting rules from a sheet."""
        cf_rules = []
        
        # Access conditional formatting through the sheet's conditional_formatting
        if hasattr(sheet, 'conditional_formatting'):
            for cf_range, rules in sheet.conditional_formatting._cf_rules.items():
                for rule in rules:
                    cf_rules.append({
                        'range': str(cf_range),
                        'type': type(rule).__name__,
                        'rule_details': self._describe_cf_rule(rule)
                    })
        
        return cf_rules
    
    def _describe_cf_rule(self, rule) -> Dict[str, Any]:
        """Describe a conditional formatting rule in detail."""
        description = {
            'type': type(rule).__name__
        }
        
        # Handle different rule types
        if isinstance(rule, ColorScaleRule):
            description.update({
                'start_type': rule.start_type,
                'start_value': rule.start_value,
                'start_color': self._get_color_description(rule.start_color),
                'end_type': rule.end_type,
                'end_value': rule.end_value,
                'end_color': self._get_color_description(rule.end_color)
            })
        elif isinstance(rule, DataBarRule):
            description.update({
                'min_type': rule.min_type,
                'min_value': rule.min_value,
                'max_type': rule.max_type,
                'max_value': rule.max_value,
                'color': self._get_color_description(rule.color)
            })
        elif isinstance(rule, IconSetRule):
            description.update({
                'icon_style': rule.iconSet,
                'show_value': rule.showValue,
                'percent': rule.percent,
                'reverse': rule.reverse
            })
        else:
            # Generic rule properties
            if hasattr(rule, 'formula'):
                description['formula'] = rule.formula
            if hasattr(rule, 'text'):
                description['text'] = rule.text
            if hasattr(rule, 'operator'):
                description['operator'] = rule.operator
        
        return description
    
    def _compare_cf_rules(self, rules_a: List[Dict], rules_b: List[Dict]) -> Dict[str, Any]:
        """Compare conditional formatting rules between two sheets."""
        
        # Simple comparison - could be enhanced with more sophisticated matching
        return {
            'total_rules_a': len(rules_a),
            'total_rules_b': len(rules_b),
            'rule_difference': len(rules_b) - len(rules_a),
            'rules_added': max(0, len(rules_b) - len(rules_a)),
            'rules_removed': max(0, len(rules_a) - len(rules_b)),
            'detailed_comparison': 'Advanced rule-by-rule comparison available on request'
        }
    
    def _compare_sheet_structure(self, sheet_a, sheet_b) -> Dict[str, Any]:
        """Compare structural elements between two sheets."""
        
        return {
            'merged_cells': self._compare_merged_cells(sheet_a, sheet_b),
            'column_dimensions': self._compare_column_dimensions(sheet_a, sheet_b),
            'row_dimensions': self._compare_row_dimensions(sheet_a, sheet_b),
            'sheet_properties': self._compare_sheet_properties(sheet_a, sheet_b),
            'protection_settings': self._compare_sheet_protection(sheet_a, sheet_b)
        }
    
    def _compare_merged_cells(self, sheet_a, sheet_b) -> Dict[str, Any]:
        """Compare merged cell ranges between two sheets."""
        
        merged_a = set(str(merged_range) for merged_range in sheet_a.merged_cells.ranges)
        merged_b = set(str(merged_range) for merged_range in sheet_b.merged_cells.ranges)
        
        return {
            'sheet_a_merged': len(merged_a),
            'sheet_b_merged': len(merged_b),
            'common_merged': len(merged_a & merged_b),
            'only_in_a': list(merged_a - merged_b),
            'only_in_b': list(merged_b - merged_a),
            'differences_count': len(merged_a.symmetric_difference(merged_b))
        }
    
    def _compare_column_dimensions(self, sheet_a, sheet_b) -> Dict[str, Any]:
        """Compare column widths and properties."""
        
        col_diff = []
        all_cols = set(sheet_a.column_dimensions.keys()) | set(sheet_b.column_dimensions.keys())
        
        for col in all_cols:
            dim_a = sheet_a.column_dimensions.get(col)
            dim_b = sheet_b.column_dimensions.get(col)
            
            if dim_a and dim_b:
                if dim_a.width != dim_b.width:
                    col_diff.append({
                        'column': col,
                        'width_a': dim_a.width,
                        'width_b': dim_b.width,
                        'change': 'width_changed'
                    })
                
                if dim_a.hidden != dim_b.hidden:
                    col_diff.append({
                        'column': col,
                        'hidden_a': dim_a.hidden,
                        'hidden_b': dim_b.hidden,
                        'change': 'visibility_changed'
                    })
            
            elif dim_a and not dim_b:
                col_diff.append({
                    'column': col,
                    'change': 'dimension_removed_in_b'
                })
            elif dim_b and not dim_a:
                col_diff.append({
                    'column': col,
                    'change': 'dimension_added_in_b'
                })
        
        return {
            'total_differences': len(col_diff),
            'differences': col_diff
        }
    
    def _compare_row_dimensions(self, sheet_a, sheet_b) -> Dict[str, Any]:
        """Compare row heights and properties."""
        
        row_diff = []
        all_rows = set(sheet_a.row_dimensions.keys()) | set(sheet_b.row_dimensions.keys())
        
        for row in all_rows:
            dim_a = sheet_a.row_dimensions.get(row)
            dim_b = sheet_b.row_dimensions.get(row)
            
            if dim_a and dim_b:
                if dim_a.height != dim_b.height:
                    row_diff.append({
                        'row': row,
                        'height_a': dim_a.height,
                        'height_b': dim_b.height,
                        'change': 'height_changed'
                    })
                
                if dim_a.hidden != dim_b.hidden:
                    row_diff.append({
                        'row': row,
                        'hidden_a': dim_a.hidden,
                        'hidden_b': dim_b.hidden,
                        'change': 'visibility_changed'
                    })
            
            elif dim_a and not dim_b:
                row_diff.append({
                    'row': row,
                    'change': 'dimension_removed_in_b'
                })
            elif dim_b and not dim_a:
                row_diff.append({
                    'row': row,
                    'change': 'dimension_added_in_b'
                })
        
        return {
            'total_differences': len(row_diff),
            'differences': row_diff
        }
    
    def _compare_sheet_properties(self, sheet_a, sheet_b) -> Dict[str, Any]:
        """Compare general sheet properties."""
        
        properties_diff = []
        
        # Compare sheet titles
        if sheet_a.title != sheet_b.title:
            properties_diff.append({
                'property': 'title',
                'value_a': sheet_a.title,
                'value_b': sheet_b.title
            })
        
        # Compare sheet visibility
        if hasattr(sheet_a, 'sheet_state') and hasattr(sheet_b, 'sheet_state'):
            if sheet_a.sheet_state != sheet_b.sheet_state:
                properties_diff.append({
                    'property': 'visibility',
                    'value_a': sheet_a.sheet_state,
                    'value_b': sheet_b.sheet_state
                })
        
        # Compare tab colors
        if hasattr(sheet_a, 'sheet_properties') and hasattr(sheet_b, 'sheet_properties'):
            if sheet_a.sheet_properties.tabColor != sheet_b.sheet_properties.tabColor:
                properties_diff.append({
                    'property': 'tab_color',
                    'value_a': self._get_color_description(sheet_a.sheet_properties.tabColor),
                    'value_b': self._get_color_description(sheet_b.sheet_properties.tabColor)
                })
        
        return {
            'differences_count': len(properties_diff),
            'differences': properties_diff
        }
    
    def _compare_sheet_protection(self, sheet_a, sheet_b) -> Dict[str, Any]:
        """Compare sheet protection settings."""
        
        protection_diff = []
        
        if hasattr(sheet_a, 'protection') and hasattr(sheet_b, 'protection'):
            prot_a = sheet_a.protection
            prot_b = sheet_b.protection
            
            # Compare protection attributes
            protection_attrs = [
                'sheet', 'objects', 'scenarios', 'formatCells', 'formatColumns',
                'formatRows', 'insertColumns', 'insertRows', 'insertHyperlinks',
                'deleteColumns', 'deleteRows', 'selectLockedCells', 'sort',
                'autoFilter', 'pivotTables', 'selectUnlockedCells'
            ]
            
            for attr in protection_attrs:
                val_a = getattr(prot_a, attr, None)
                val_b = getattr(prot_b, attr, None)
                
                if val_a != val_b:
                    protection_diff.append({
                        'setting': attr,
                        'value_a': val_a,
                        'value_b': val_b
                    })
        
        return {
            'differences_count': len(protection_diff),
            'differences': protection_diff
        }
    
    def _calculate_formatting_statistics(self, sheet_a, sheet_b) -> Dict[str, Any]:
        """Calculate overall formatting statistics."""
        
        # Get all formatted cells
        cells_a = self._get_formatted_cells(sheet_a)
        cells_b = self._get_formatted_cells(sheet_b)
        
        # Count different types of formatting
        font_formatted_a = sum(1 for cell in cells_a.values() if self._has_font_formatting(cell))
        font_formatted_b = sum(1 for cell in cells_b.values() if self._has_font_formatting(cell))
        
        fill_formatted_a = sum(1 for cell in cells_a.values() if self._has_fill_formatting(cell))
        fill_formatted_b = sum(1 for cell in cells_b.values() if self._has_fill_formatting(cell))
        
        border_formatted_a = sum(1 for cell in cells_a.values() if self._has_border_formatting(cell))
        border_formatted_b = sum(1 for cell in cells_b.values() if self._has_border_formatting(cell))
        
        return {
            'overview': {
                'total_formatted_cells_a': len(cells_a),
                'total_formatted_cells_b': len(cells_b),
                'formatting_coverage_change': len(cells_b) - len(cells_a)
            },
            'formatting_types': {
                'font_formatting': {
                    'sheet_a': font_formatted_a,
                    'sheet_b': font_formatted_b,
                    'change': font_formatted_b - font_formatted_a
                },
                'fill_formatting': {
                    'sheet_a': fill_formatted_a,
                    'sheet_b': fill_formatted_b,
                    'change': fill_formatted_b - fill_formatted_a
                },
                'border_formatting': {
                    'sheet_a': border_formatted_a,
                    'sheet_b': border_formatted_b,
                    'change': border_formatted_b - border_formatted_a
                }
            },
            'complexity_metrics': {
                'merged_cells_a': len(list(sheet_a.merged_cells.ranges)),
                'merged_cells_b': len(list(sheet_b.merged_cells.ranges)),
                'conditional_formatting_a': len(self._extract_conditional_formatting(sheet_a)),
                'conditional_formatting_b': len(self._extract_conditional_formatting(sheet_b))
            }
        }
    
    def _has_font_formatting(self, cell) -> bool:
        """Check if cell has custom font formatting."""
        default_font = Font()
        return cell.font != default_font
    
    def _has_fill_formatting(self, cell) -> bool:
        """Check if cell has custom fill formatting."""
        default_fill = Fill()
        return cell.fill != default_fill
    
    def _has_border_formatting(self, cell) -> bool:
        """Check if cell has custom border formatting."""
        default_border = Border()
        return cell.border != default_border
    
    def _generate_formatting_diff_report(self, sheet_a, sheet_b) -> Dict[str, Any]:
        """Generate a comprehensive formatting difference report."""
        
        # Get sample of most significant differences
        cells_a = self._get_formatted_cells(sheet_a)
        cells_b = self._get_formatted_cells(sheet_b)
        all_coords = set(cells_a.keys()) | set(cells_b.keys())
        
        significant_differences = []
        
        # Find the most significant formatting changes
        for coord in sorted(all_coords)[:50]:  # Limit to first 50 for performance
            cell_a = cells_a.get(coord)
            cell_b = cells_b.get(coord)
            
            diff = self._compare_cell_formats(coord, cell_a, cell_b)
            if diff['has_differences']:
                significance_score = self._calculate_difference_significance(diff['differences'])
                if significance_score > 3:  # Threshold for significance
                    significant_differences.append({
                        'cell': coord,
                        'significance_score': significance_score,
                        'differences': diff['differences']
                    })
        
        # Sort by significance
        significant_differences.sort(key=lambda x: x['significance_score'], reverse=True)
        
        return {
            'most_significant_changes': significant_differences[:20],
            'total_significant_changes': len(significant_differences),
            'change_impact_assessment': self._assess_change_impact(significant_differences)
        }
    
    def _calculate_difference_significance(self, differences: Dict[str, List]) -> int:
        """Calculate a significance score for formatting differences."""
        score = 0
        
        # Weight different types of changes
        weights = {
            'font': 2,
            'fill': 3,
            'border': 2,
            'alignment': 1,
            'number_format': 2,
            'protection': 1
        }
        
        for category, changes in differences.items():
            if changes:
                score += len(changes) * weights.get(category, 1)
        
        return score
    
    def _assess_change_impact(self, significant_differences: List[Dict]) -> Dict[str, Any]:
        """Assess the overall impact of formatting changes."""
        
        if not significant_differences:
            return {
                'impact_level': 'None',
                'summary': 'No significant formatting changes detected'
            }
        
        total_changes = len(significant_differences)
        avg_significance = sum(diff['significance_score'] for diff in significant_differences) / total_changes
        
        if avg_significance > 8:
            impact_level = 'High'
            summary = f"Major formatting changes detected ({total_changes} significant changes)"
        elif avg_significance > 5:
            impact_level = 'Medium'
            summary = f"Moderate formatting changes detected ({total_changes} changes)"
        else:
            impact_level = 'Low'
            summary = f"Minor formatting changes detected ({total_changes} changes)"
        
        return {
            'impact_level': impact_level,
            'summary': summary,
            'total_significant_changes': total_changes,
            'average_significance_score': avg_significance
        }
    
    def _summarize_formatting_differences(self, format_categories: Dict[str, List]) -> Dict[str, Any]:
        """Summarize formatting differences by category."""
        
        summary = {}
        
        for category, differences in format_categories.items():
            category_name = category.replace('_differences', '').replace('_', ' ').title()
            summary[category_name] = {
                'count': len(differences),
                'percentage': 0,  # Will be calculated if needed
                'examples': [diff['cell'] for diff in differences[:5]]  # First 5 examples
            }
        
        return summary

# Convenience function for easy importing
def compare_excel_formatting(file_a: Any, file_b: Any, 
                            sheet_name_a: Optional[str] = None, 
                            sheet_name_b: Optional[str] = None) -> Dict[str, Any]:
    """
    Convenience function to compare Excel formatting between two files.
    
    Args:
        file_a: First Excel file
        file_b: Second Excel file  
        sheet_name_a: Sheet name in first file (optional)
        sheet_name_b: Sheet name in second file (optional)
    
    Returns:
        Comprehensive formatting comparison results
    """
    comparator = FormattingComparator()
    return comparator.compare_excel_formatting(file_a, file_b, sheet_name_a, sheet_name_b)