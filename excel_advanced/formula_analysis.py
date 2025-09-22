# Formula Analysis Module for Excel Files
# Provides comprehensive formula extraction, validation, and dependency analysis

import pandas as pd
import numpy as np
from typing import Dict, List, Set, Tuple, Optional, Any
import re
import streamlit as st
from datetime import datetime
import networkx as nx

# Handle openpyxl import gracefully
try:
    from openpyxl import load_workbook
    from openpyxl.cell import Cell
    from openpyxl.formula.translate import Translator
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not available. Formula analysis will be limited.")

class FormulaAnalyzer:
    """
    Comprehensive Excel formula analysis tool
    
    Features:
    - Formula extraction from Excel cells
    - Formula comparison between sheets
    - Broken formula detection
    - Circular reference identification
    - Dependency mapping and impact analysis
    """
    
    def __init__(self):
        self.formulas = {}
        self.dependencies = nx.DiGraph()  # Directed graph for dependencies
        self.circular_references = []
        self.broken_formulas = []
        self.formula_stats = {}
        
    def analyze_excel_formulas(self, uploaded_file, sheet_names: List[str] = None) -> Dict[str, Any]:
        """
        Main method to analyze formulas in Excel file
        
        Args:
            uploaded_file: Streamlit uploaded file object
            sheet_names: List of sheet names to analyze (None for all)
            
        Returns:
            Comprehensive formula analysis results
        """
        
        if not OPENPYXL_AVAILABLE:
            return {
                'error': 'openpyxl library not available. Please install with: pip install openpyxl',
                'status': 'failed'
            }
        
        try:
            # Load workbook with formulas (data_only=False)
            wb = load_workbook(uploaded_file, data_only=False)
            
            if sheet_names is None:
                sheet_names = wb.sheetnames
            
            # Initialize analysis results
            analysis_results = {
                'file_info': {
                    'total_sheets': len(wb.sheetnames),
                    'analyzed_sheets': len(sheet_names),
                    'sheet_names': sheet_names
                },
                'formula_extraction': {},
                'formula_validation': {},
                'dependency_analysis': {},
                'summary_statistics': {},
                'status': 'success'
            }
            
            # Extract formulas from each sheet
            total_formulas = 0
            for sheet_name in sheet_names:
                if sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    formulas = self._extract_sheet_formulas(sheet, sheet_name)
                    analysis_results['formula_extraction'][sheet_name] = formulas
                    total_formulas += len(formulas)
            
            # Validate formulas
            validation_results = self._validate_formulas(analysis_results['formula_extraction'])
            analysis_results['formula_validation'] = validation_results
            
            # Build dependency graph
            dependency_results = self._build_dependency_graph(analysis_results['formula_extraction'])
            analysis_results['dependency_analysis'] = dependency_results
            
            # Generate summary statistics
            analysis_results['summary_statistics'] = self._generate_formula_statistics(
                analysis_results['formula_extraction'],
                validation_results,
                dependency_results
            )
            
            return analysis_results
            
        except Exception as e:
            return {
                'error': f'Error analyzing formulas: {str(e)}',
                'status': 'failed'
            }
    
    def _extract_sheet_formulas(self, sheet, sheet_name: str) -> Dict[str, Any]:
        """Extract all formulas from a worksheet"""
        
        formulas = {
            'cell_formulas': {},
            'formula_types': {},
            'complex_formulas': [],
            'simple_formulas': [],
            'array_formulas': [],
            'total_count': 0
        }
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f' and cell.coordinate:  # Formula cell
                    formula_text = str(cell.value) if cell.value else ''
                    
                    if formula_text.startswith('='):
                        cell_ref = cell.coordinate
                        
                        # Store formula details
                        formula_info = {
                            'formula': formula_text,
                            'sheet': sheet_name,
                            'cell': cell_ref,
                            'row': cell.row,
                            'column': cell.column,
                            'complexity': self._calculate_formula_complexity(formula_text),
                            'functions_used': self._extract_functions(formula_text),
                            'cell_references': self._extract_cell_references(formula_text),
                            'external_references': self._extract_external_references(formula_text)
                        }
                        
                        formulas['cell_formulas'][cell_ref] = formula_info
                        
                        # Categorize formula
                        complexity = formula_info['complexity']
                        if complexity > 10:
                            formulas['complex_formulas'].append(formula_info)
                        else:
                            formulas['simple_formulas'].append(formula_info)
                        
                        # Check for array formulas
                        if self._is_array_formula(formula_text):
                            formulas['array_formulas'].append(formula_info)
                        
                        # Categorize by function type
                        main_function = self._get_main_function(formula_text)
                        if main_function not in formulas['formula_types']:
                            formulas['formula_types'][main_function] = []
                        formulas['formula_types'][main_function].append(formula_info)
        
        formulas['total_count'] = len(formulas['cell_formulas'])
        return formulas
    
    def _calculate_formula_complexity(self, formula: str) -> int:
        """Calculate complexity score for a formula"""
        
        complexity = 0
        
        # Base complexity
        complexity += len(formula) // 10
        
        # Function complexity
        functions = self._extract_functions(formula)
        complexity += len(functions) * 2
        
        # Nested functions (count parentheses depth)
        max_depth = 0
        current_depth = 0
        for char in formula:
            if char == '(':
                current_depth += 1
                max_depth = max(max_depth, current_depth)
            elif char == ')':
                current_depth -= 1
        complexity += max_depth * 3
        
        # Cell references
        cell_refs = self._extract_cell_references(formula)
        complexity += len(cell_refs)
        
        # Operators
        operators = ['+', '-', '*', '/', '^', '&', '>', '<', '=']
        for op in operators:
            complexity += formula.count(op)
        
        return complexity
    
    def _extract_functions(self, formula: str) -> List[str]:
        """Extract all Excel functions from a formula"""
        
        # Common Excel functions pattern
        function_pattern = r'([A-Z][A-Z0-9]*)\s*\('
        functions = re.findall(function_pattern, formula.upper())
        
        # Remove duplicates while preserving order
        unique_functions = []
        for func in functions:
            if func not in unique_functions:
                unique_functions.append(func)
        
        return unique_functions
    
    def _extract_cell_references(self, formula: str) -> List[str]:
        """Extract all cell references from a formula"""
        
        # Pattern for cell references (A1, $A$1, Sheet1!A1, etc.)
        cell_pattern = r'(?:([A-Za-z_][A-Za-z0-9_]*!)?)(\$?[A-Z]+\$?[0-9]+)'
        matches = re.findall(cell_pattern, formula)
        
        cell_refs = []
        for sheet_ref, cell_ref in matches:
            full_ref = f"{sheet_ref}{cell_ref}" if sheet_ref else cell_ref
            cell_refs.append(full_ref)
        
        return cell_refs
    
    def _extract_external_references(self, formula: str) -> List[str]:
        """Extract external file references from a formula"""
        
        # Pattern for external references [Workbook.xlsx]Sheet!A1
        external_pattern = r'\[([^\]]+)\]([^!]+)!'
        matches = re.findall(external_pattern, formula)
        
        external_refs = []
        for workbook, sheet in matches:
            external_refs.append(f"[{workbook}]{sheet}")
        
        return external_refs
    
    def _is_array_formula(self, formula: str) -> bool:
        """Check if formula is an array formula"""
        
        # Array formulas typically contain array functions or array constants
        array_functions = ['TRANSPOSE', 'MMULT', 'INDEX', 'MATCH', 'SUMPRODUCT']
        array_patterns = ['{', '}', '=', '<=', '>=', '<>']
        
        formula_upper = formula.upper()
        
        # Check for array functions
        for func in array_functions:
            if func in formula_upper:
                return True
        
        # Check for array constants
        if '{' in formula and '}' in formula:
            return True
        
        return False
    
    def _get_main_function(self, formula: str) -> str:
        """Get the primary function in a formula"""
        
        functions = self._extract_functions(formula)
        if functions:
            return functions[0]  # First function is usually the main one
        else:
            return 'SIMPLE'  # Simple arithmetic or reference
    
    def _validate_formulas(self, formula_data: Dict[str, Any]) -> Dict[str, Any]:
        """Validate formulas for errors and issues"""
        
        validation_results = {
            'broken_formulas': [],
            'circular_references': [],
            'missing_references': [],
            'external_dependencies': [],
            'validation_summary': {}
        }
        
        all_cell_refs = set()
        formula_cells = {}
        
        # Collect all cell references and formulas
        for sheet_name, sheet_data in formula_data.items():
            for cell_ref, formula_info in sheet_data['cell_formulas'].items():
                full_cell_ref = f"{sheet_name}!{cell_ref}"
                all_cell_refs.add(full_cell_ref)
                formula_cells[full_cell_ref] = formula_info
        
        # Validate each formula
        for full_cell_ref, formula_info in formula_cells.items():
            formula = formula_info['formula']
            
            # Check for missing references
            for ref in formula_info['cell_references']:
                if '!' not in ref:  # Same sheet reference
                    ref = f"{formula_info['sheet']}!{ref}"
                
                if ref not in all_cell_refs:
                    validation_results['missing_references'].append({
                        'formula_cell': full_cell_ref,
                        'missing_reference': ref,
                        'formula': formula
                    })
            
            # Check for external dependencies
            if formula_info['external_references']:
                validation_results['external_dependencies'].append({
                    'formula_cell': full_cell_ref,
                    'external_refs': formula_info['external_references'],
                    'formula': formula
                })
            
            # Check for potential circular references (simplified)
            if self._has_potential_circular_reference(formula_info, formula_cells):
                validation_results['circular_references'].append({
                    'formula_cell': full_cell_ref,
                    'formula': formula,
                    'issue': 'Potential circular reference detected'
                })
        
        # Generate validation summary
        validation_results['validation_summary'] = {
            'total_formulas_validated': len(formula_cells),
            'broken_formulas_count': len(validation_results['broken_formulas']),
            'circular_references_count': len(validation_results['circular_references']),
            'missing_references_count': len(validation_results['missing_references']),
            'external_dependencies_count': len(validation_results['external_dependencies']),
            'validation_score': self._calculate_validation_score(validation_results, len(formula_cells))
        }
        
        return validation_results
    
    def _has_potential_circular_reference(self, formula_info: Dict, all_formulas: Dict) -> bool:
        """Check for potential circular references (simplified detection)"""
        
        # This is a simplified circular reference detection
        # A more complete implementation would build a full dependency graph
        
        formula_cell = f"{formula_info['sheet']}!{formula_info['cell']}"
        
        for cell_ref in formula_info['cell_references']:
            if '!' not in cell_ref:
                cell_ref = f"{formula_info['sheet']}!{cell_ref}"
            
            # Check if any referenced cell refers back to this cell
            if cell_ref in all_formulas:
                ref_formula = all_formulas[cell_ref]
                for ref_cell_ref in ref_formula['cell_references']:
                    if '!' not in ref_cell_ref:
                        ref_cell_ref = f"{ref_formula['sheet']}!{ref_cell_ref}"
                    
                    if ref_cell_ref == formula_cell:
                        return True
        
        return False
    
    def _calculate_validation_score(self, validation_results: Dict, total_formulas: int) -> float:
        """Calculate overall formula validation score"""
        
        if total_formulas == 0:
            return 100.0
        
        issues = (
            len(validation_results['broken_formulas']) +
            len(validation_results['circular_references']) +
            len(validation_results['missing_references'])
        )
        
        score = max(0, 100 - (issues / total_formulas * 100))
        return round(score, 1)
    
    def _build_dependency_graph(self, formula_data: Dict[str, Any]) -> Dict[str, Any]:
        """Build formula dependency graph and analyze relationships"""
        
        dependency_results = {
            'dependency_tree': {},
            'precedents': {},  # Cells that this cell depends on
            'dependents': {},  # Cells that depend on this cell
            'independent_cells': [],
            'highly_connected_cells': [],
            'dependency_levels': {},
            'impact_analysis': {}
        }
        
        # Build dependency relationships
        all_cells = {}
        
        for sheet_name, sheet_data in formula_data.items():
            for cell_ref, formula_info in sheet_data['cell_formulas'].items():
                full_cell_ref = f"{sheet_name}!{cell_ref}"
                all_cells[full_cell_ref] = formula_info
                
                # Initialize precedents and dependents
                dependency_results['precedents'][full_cell_ref] = []
                dependency_results['dependents'][full_cell_ref] = []
        
        # Analyze dependencies
        for full_cell_ref, formula_info in all_cells.items():
            precedents = []
            
            for cell_ref in formula_info['cell_references']:
                if '!' not in cell_ref:  # Same sheet reference
                    precedent = f"{formula_info['sheet']}!{cell_ref}"
                else:
                    precedent = cell_ref
                
                precedents.append(precedent)
                
                # Add to dependents of the referenced cell
                if precedent in dependency_results['dependents']:
                    dependency_results['dependents'][precedent].append(full_cell_ref)
            
            dependency_results['precedents'][full_cell_ref] = precedents
        
        # Find independent cells (no precedents)
        for cell_ref, precedents in dependency_results['precedents'].items():
            if not precedents:
                dependency_results['independent_cells'].append(cell_ref)
        
        # Find highly connected cells (many dependents)
        for cell_ref, dependents in dependency_results['dependents'].items():
            if len(dependents) >= 5:  # Threshold for "highly connected"
                dependency_results['highly_connected_cells'].append({
                    'cell': cell_ref,
                    'dependent_count': len(dependents),
                    'dependents': dependents
                })
        
        # Calculate dependency levels (depth from independent cells)
        dependency_results['dependency_levels'] = self._calculate_dependency_levels(
            dependency_results['precedents'],
            dependency_results['independent_cells']
        )
        
        # Perform impact analysis
        dependency_results['impact_analysis'] = self._perform_impact_analysis(
            dependency_results['dependents'],
            dependency_results['highly_connected_cells']
        )
        
        return dependency_results
    
    def _calculate_dependency_levels(self, precedents: Dict, independent_cells: List[str]) -> Dict[str, int]:
        """Calculate the dependency level (depth) for each cell"""
        
        levels = {}
        
        # Independent cells are at level 0
        for cell in independent_cells:
            levels[cell] = 0
        
        # Calculate levels iteratively
        changed = True
        max_iterations = 100  # Prevent infinite loops
        iteration = 0
        
        while changed and iteration < max_iterations:
            changed = False
            iteration += 1
            
            for cell, cell_precedents in precedents.items():
                if cell not in levels:
                    # Check if all precedents have levels assigned
                    precedent_levels = []
                    all_precedents_assigned = True
                    
                    for precedent in cell_precedents:
                        if precedent in levels:
                            precedent_levels.append(levels[precedent])
                        else:
                            all_precedents_assigned = False
                            break
                    
                    if all_precedents_assigned:
                        if precedent_levels:
                            levels[cell] = max(precedent_levels) + 1
                        else:
                            levels[cell] = 0
                        changed = True
        
        return levels
    
    def _perform_impact_analysis(self, dependents: Dict, highly_connected: List[Dict]) -> Dict[str, Any]:
        """Analyze the potential impact of changes to cells"""
        
        impact_analysis = {
            'high_impact_cells': [],
            'cascade_effects': {},
            'risk_assessment': {}
        }
        
        # Identify high-impact cells
        for cell_data in highly_connected:
            cell = cell_data['cell']
            dependent_count = cell_data['dependent_count']
            
            # Calculate cascade effect
            cascade_size = self._calculate_cascade_size(cell, dependents, set())
            
            impact_analysis['high_impact_cells'].append({
                'cell': cell,
                'direct_dependents': dependent_count,
                'total_cascade_size': cascade_size,
                'impact_level': 'HIGH' if cascade_size > 10 else 'MEDIUM' if cascade_size > 5 else 'LOW'
            })
        
        # Generate risk assessment
        impact_analysis['risk_assessment'] = {
            'total_high_impact_cells': len([c for c in impact_analysis['high_impact_cells'] if c['impact_level'] == 'HIGH']),
            'average_cascade_size': np.mean([c['total_cascade_size'] for c in impact_analysis['high_impact_cells']]) if impact_analysis['high_impact_cells'] else 0,
            'max_cascade_size': max([c['total_cascade_size'] for c in impact_analysis['high_impact_cells']]) if impact_analysis['high_impact_cells'] else 0
        }
        
        return impact_analysis
    
    def _calculate_cascade_size(self, cell: str, dependents: Dict, visited: Set[str]) -> int:
        """Calculate the total cascade effect size for a cell"""
        
        if cell in visited:
            return 0
        
        visited.add(cell)
        cascade_size = 0
        
        if cell in dependents:
            direct_dependents = dependents[cell]
            cascade_size += len(direct_dependents)
            
            # Recursively calculate cascade for dependents
            for dependent in direct_dependents:
                cascade_size += self._calculate_cascade_size(dependent, dependents, visited.copy())
        
        return cascade_size
    
    def _generate_formula_statistics(self, formula_data: Dict, validation_data: Dict, dependency_data: Dict) -> Dict[str, Any]:
        """Generate comprehensive formula statistics"""
        
        stats = {
            'overview': {},
            'complexity_analysis': {},
            'function_usage': {},
            'dependency_metrics': {},
            'quality_metrics': {}
        }
        
        # Overview statistics
        total_formulas = 0
        total_sheets = len(formula_data)
        
        all_functions = []
        complexity_scores = []
        
        for sheet_name, sheet_data in formula_data.items():
            sheet_formulas = len(sheet_data['cell_formulas'])
            total_formulas += sheet_formulas
            
            for cell_ref, formula_info in sheet_data['cell_formulas'].items():
                all_functions.extend(formula_info['functions_used'])
                complexity_scores.append(formula_info['complexity'])
        
        stats['overview'] = {
            'total_formulas': total_formulas,
            'total_sheets_with_formulas': total_sheets,
            'average_formulas_per_sheet': round(total_formulas / total_sheets if total_sheets > 0 else 0, 1),
            'unique_functions_used': len(set(all_functions)),
            'most_complex_formula_score': max(complexity_scores) if complexity_scores else 0
        }
        
        # Complexity analysis
        if complexity_scores:
            stats['complexity_analysis'] = {
                'average_complexity': round(np.mean(complexity_scores), 1),
                'median_complexity': round(np.median(complexity_scores), 1),
                'complexity_std': round(np.std(complexity_scores), 1),
                'simple_formulas': len([s for s in complexity_scores if s <= 5]),
                'moderate_formulas': len([s for s in complexity_scores if 5 < s <= 15]),
                'complex_formulas': len([s for s in complexity_scores if s > 15])
            }
        
        # Function usage analysis
        function_counts = {}
        for func in all_functions:
            function_counts[func] = function_counts.get(func, 0) + 1
        
        # Sort by usage frequency
        sorted_functions = sorted(function_counts.items(), key=lambda x: x[1], reverse=True)
        
        stats['function_usage'] = {
            'total_function_calls': len(all_functions),
            'top_10_functions': sorted_functions[:10],
            'function_diversity_score': round(len(set(all_functions)) / len(all_functions) * 100 if all_functions else 0, 1)
        }
        
        # Dependency metrics
        stats['dependency_metrics'] = {
            'independent_cells': len(dependency_data.get('independent_cells', [])),
            'highly_connected_cells': len(dependency_data.get('highly_connected_cells', [])),
            'max_dependency_level': max(dependency_data.get('dependency_levels', {}).values()) if dependency_data.get('dependency_levels') else 0,
            'average_precedents': round(np.mean([len(prec) for prec in dependency_data.get('precedents', {}).values()]) if dependency_data.get('precedents') else 0, 1)
        }
        
        # Quality metrics
        validation_summary = validation_data.get('validation_summary', {})
        stats['quality_metrics'] = {
            'validation_score': validation_summary.get('validation_score', 0),
            'error_rate': round((validation_summary.get('broken_formulas_count', 0) + 
                               validation_summary.get('missing_references_count', 0)) / total_formulas * 100 if total_formulas > 0 else 0, 1),
            'external_dependency_rate': round(validation_summary.get('external_dependencies_count', 0) / total_formulas * 100 if total_formulas > 0 else 0, 1),
            'circular_reference_risk': 'HIGH' if validation_summary.get('circular_references_count', 0) > 0 else 'LOW'
        }
        
        return stats
    
    def compare_formulas_between_sheets(self, analysis_a: Dict, analysis_b: Dict, sheet_a: str, sheet_b: str) -> Dict[str, Any]:
        """Compare formulas between two sheets"""
        
        comparison_results = {
            'summary': {},
            'formula_differences': [],
            'function_usage_comparison': {},
            'complexity_comparison': {},
            'dependency_comparison': {},
            'recommendations': []
        }
        
        if sheet_a not in analysis_a['formula_extraction'] or sheet_b not in analysis_b['formula_extraction']:
            comparison_results['error'] = 'One or both sheets not found in analysis results'
            return comparison_results
        
        formulas_a = analysis_a['formula_extraction'][sheet_a]['cell_formulas']
        formulas_b = analysis_b['formula_extraction'][sheet_b]['cell_formulas']
        
        # Summary comparison
        comparison_results['summary'] = {
            'formulas_count_a': len(formulas_a),
            'formulas_count_b': len(formulas_b),
            'common_cells': len(set(formulas_a.keys()) & set(formulas_b.keys())),
            'unique_to_a': len(set(formulas_a.keys()) - set(formulas_b.keys())),
            'unique_to_b': len(set(formulas_b.keys()) - set(formulas_a.keys()))
        }
        
        # Compare formulas in common cells
        common_cells = set(formulas_a.keys()) & set(formulas_b.keys())
        
        for cell_ref in common_cells:
            formula_a = formulas_a[cell_ref]['formula']
            formula_b = formulas_b[cell_ref]['formula']
            
            if formula_a != formula_b:
                comparison_results['formula_differences'].append({
                    'cell': cell_ref,
                    'formula_a': formula_a,
                    'formula_b': formula_b,
                    'difference_type': self._categorize_formula_difference(formula_a, formula_b)
                })
        
        # Function usage comparison
        functions_a = []
        functions_b = []
        
        for formula_info in formulas_a.values():
            functions_a.extend(formula_info['functions_used'])
        
        for formula_info in formulas_b.values():
            functions_b.extend(formula_info['functions_used'])
        
        unique_functions_a = set(functions_a)
        unique_functions_b = set(functions_b)
        
        comparison_results['function_usage_comparison'] = {
            'common_functions': list(unique_functions_a & unique_functions_b),
            'unique_to_a': list(unique_functions_a - unique_functions_b),
            'unique_to_b': list(unique_functions_b - unique_functions_a),
            'function_overlap_percentage': round(len(unique_functions_a & unique_functions_b) / len(unique_functions_a | unique_functions_b) * 100 if unique_functions_a | unique_functions_b else 0, 1)
        }
        
        # Generate recommendations
        comparison_results['recommendations'] = self._generate_formula_comparison_recommendations(comparison_results)
        
        return comparison_results
    
    def _categorize_formula_difference(self, formula_a: str, formula_b: str) -> str:
        """Categorize the type of difference between two formulas"""
        
        # Simple categorization logic
        if len(formula_a) != len(formula_b):
            if abs(len(formula_a) - len(formula_b)) > 20:
                return 'MAJOR_STRUCTURAL_CHANGE'
            else:
                return 'MINOR_MODIFICATION'
        
        # Check if only cell references changed
        functions_a = self._extract_functions(formula_a)
        functions_b = self._extract_functions(formula_b)
        
        if functions_a == functions_b:
            return 'REFERENCE_CHANGE'
        else:
            return 'FUNCTION_CHANGE'
    
    def _generate_formula_comparison_recommendations(self, comparison_results: Dict) -> List[str]:
        """Generate recommendations based on formula comparison"""
        
        recommendations = []
        
        summary = comparison_results['summary']
        
        if summary['unique_to_a'] > summary['unique_to_b']:
            recommendations.append(f"Sheet A has {summary['unique_to_a']} more formulas than Sheet B. Consider if these are necessary.")
        
        if len(comparison_results['formula_differences']) > 0:
            recommendations.append(f"Found {len(comparison_results['formula_differences'])} formula differences in common cells. Review for consistency.")
        
        function_comparison = comparison_results['function_usage_comparison']
        if function_comparison['function_overlap_percentage'] < 70:
            recommendations.append("Low function overlap between sheets suggests different calculation approaches. Consider standardization.")
        
        return recommendations

# Export the class
__all__ = ['FormulaAnalyzer']